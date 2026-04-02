from datetime import datetime
from io import BytesIO
from unittest.mock import MagicMock

import openpyxl

from app.services.export_service import ExportService


def _make_mock_card(
    target_text="hello",
    target_meaning="greeting",
    card_type="phrase",
    context_sentence="Hello, how are you?",
    context_translation="Translation here",
    cloze_sentence="___, how are you?",
    tags=None,
    interval=1,
    ease_factor=2.5,
    next_review=None,
    created_at=None,
):
    card = MagicMock()
    card.type = card_type
    card.target_text = target_text
    card.target_meaning = target_meaning
    card.context_sentence = context_sentence
    card.context_translation = context_translation
    card.cloze_sentence = cloze_sentence
    card.interval = interval
    card.ease_factor = ease_factor
    card.next_review = next_review or datetime(2026, 4, 1)
    card.created_at = created_at or datetime(2026, 3, 1)
    card.tags = tags or []
    return card


def _make_mock_tag(name):
    tag = MagicMock()
    tag.name = name
    return tag


class TestExportService:
    def test_generates_xlsx_with_headers(self):
        service = ExportService()
        buf = service.cards_to_xlsx([])
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Type",
            "Word/Phrase",
            "Meaning",
            "Context Sentence",
            "Context Translation",
            "Cloze Sentence",
            "Tags",
            "Interval (days)",
            "Ease Factor",
            "Next Review",
            "Created At",
        ]

    def test_generates_xlsx_with_card_data(self):
        tag1 = _make_mock_tag("greetings")
        tag2 = _make_mock_tag("basics")
        card = _make_mock_card(tags=[tag1, tag2])
        service = ExportService()
        buf = service.cards_to_xlsx([card])
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        row = [cell.value for cell in ws[2]]
        assert row[0] == "phrase"
        assert row[1] == "hello"
        assert row[2] == "greeting"
        assert row[3] == "Hello, how are you?"
        assert row[4] == "Translation here"
        assert row[5] == "___, how are you?"
        assert row[6] == "greetings, basics"
        assert row[7] == 1
        assert row[8] == 2.5

    def test_generates_xlsx_with_multiple_cards(self):
        cards = [_make_mock_card(target_text=f"word{i}") for i in range(3)]
        service = ExportService()
        buf = service.cards_to_xlsx(cards)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 3 data rows

    def test_returns_bytes_io(self):
        service = ExportService()
        buf = service.cards_to_xlsx([])
        assert isinstance(buf, BytesIO)
        assert buf.tell() == 0  # seek position reset to start
